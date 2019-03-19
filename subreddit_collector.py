# Collects multiple subreddits posts for the post id, title, and url and saves it in an xlsx file
# Skips posts labeled NSFW
# Skips posts already saved
import praw
import openpyxl
import json

# Reddit credentials
reddit = praw.Reddit(client_id='2bwKvsxuAFExCw',
                     client_secret='bnuTQT8tyZTG-1zyhfK6Agp2BR8',
                     user_agent='android:com.Finder:v1 (by /u/BriefMath)')

# Imports json data to a list.
json_data = open('subList.txt', 'r')
subList = json.load(json_data)

# Opens workbook
wb = openpyxl.load_workbook('data.xlsx')

# Checks if every item in the json file has a unique sheet in reddit.xlsx. Will create sheet if it
# does not exist.
for sub in subList:
    if sub not in wb.sheetnames:
        wb.create_sheet(title=sub)
        print(f'Sheet created for {sub}')

# Creates a dictionary post id, title, and url for each item in subList. Writes dictionary contents to an xlsx file
# then overwrites dictionary with next Reddit entry.
for sub in subList:
    sheet = wb[sub]
    x = 1
    print('\n' + sub + '...done')
    for submission in reddit.subreddit(sub).hot(limit=100):
        sub = {}
        sub['id'] = submission.id
        sub['title'] = submission.title
        sub['url'] = submission.url
        if submission.over_18:
            continue
        elif sheet['A' + str(x)].value is None:
            sheet['A' + str(x)].value = sub['id']
            sheet['B' + str(x)].value = sub['title']
            sheet['C' + str(x)].value = sub['url']
            x = x + 1
        elif sub['id'] in sheet['A' + str(x)].value:
            continue
        elif sub['id'] not in sheet['A' + str(x)].value:
            x = x + 1

# Saves workbook
wb.save('data.xlsx')
print('\nFinished')
